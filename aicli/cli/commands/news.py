import typer
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, MofNCompleteColumn
from aicli.cli.tui import print_header, print_success, print_error, console

app = typer.Typer(help="Current affairs news parsing → filterable Excel.")


def _classify_batch(
    batch_idx: int,
    batch: list[str],
    system_prompt: str,
    client,
    model_name: str,
) -> tuple[int, list[dict], Exception | None]:
    """Worker function: sends one batch to the LLM and returns parsed records."""
    from aicli.services.news_parser import build_user_prompt, extract_json_array

    try:
        user_prompt = build_user_prompt(batch)
        response = client.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.1,
        )

        raw_output = response.choices[0].message.content or ""
        records = extract_json_array(raw_output)
        return batch_idx, records, None
    except Exception as e:
        return batch_idx, [], e


@app.command("parse")
def parse_news(
    file_path: Path = typer.Argument(
        ...,
        exists=True,
        file_okay=True,
        dir_okay=False,
        help="Path to the current affairs text/markdown file.",
    ),
    output: Path = typer.Option(
        None,
        "--output", "-o",
        help="Output Excel file path. Defaults to <input_name>_parsed.xlsx. If file exists, new rows are APPENDED.",
    ),
    batch_size: int = typer.Option(
        10,
        "--batch-size", "-b",
        help="Number of lines to send per LLM call (lower = more accurate, higher = faster).",
    ),
    month: str = typer.Option(
        "Not Specified",
        "--month", "-m",
        help="Global month to apply to all items (e.g. 'December').",
    ),
    year: str = typer.Option(
        "Not Specified",
        "--year", "-y",
        help="Global year to apply to all items (e.g. '2025').",
    ),
    workers: int = typer.Option(
        4,
        "--workers",
        "-W",
        help="Number of parallel LLM inference threads.",
        min=1,
        max=8,
    ),
    show_prompt: bool = typer.Option(
        False,
        "--show-prompt",
        help="Print the full system prompt to stdout and exit.",
    ),
):
    """
    Parse a current affairs file into a filterable Excel spreadsheet.

    Each line in the file is treated as one news item. The local LLM classifies
    each into Topic and Tags. Month and Year are supplied via CLI options.
    The original line is preserved completely as the News block.

    Supports APPEND mode: run on multiple files with the same --output and all
    rows accumulate into a single master Excel sheet.
    """
    from openai import OpenAI
    from aicli.config import config
    from aicli.services.news_parser import (
        build_system_prompt,
        filter_news_lines,
        write_excel,
        STANDARD_TOPICS,
    )

    system_prompt = build_system_prompt()

    # ── Show prompt mode ──────────────────────────────────────────────
    if show_prompt:
        console.print(system_prompt)
        raise typer.Exit()

    print_header(f"Parsing Current Affairs: {file_path.name}")

    # ── Read and filter lines ─────────────────────────────────────────
    lines = filter_news_lines(file_path)
    if not lines:
        print_error("No valid news lines found in the file.")
        raise typer.Exit(code=1)

    console.print(f"[cyan]Found {len(lines)} news lines to classify[/cyan]")
    console.print(f"[dim]Topics: {len(STANDARD_TOPICS)} · Config: Month={month}, Year={year}[/dim]")

    # ── Output path ───────────────────────────────────────────────────
    if output is None:
        output = file_path.parent / f"{file_path.stem}_parsed.xlsx"

    if output.exists():
        console.print(f"[yellow]⚡ Appending to existing file:[/yellow] {output}")
    else:
        console.print(f"[green]Creating new file:[/green] {output}")

    # ── LLM client (reuses existing LM Studio config) ─────────────────
    client = OpenAI(
        base_url=config.lm_studio_base_url,
        api_key=config.lm_studio_api_key,
    )

    # ── Split into batches ────────────────────────────────────────────
    batches = [lines[i : i + batch_size] for i in range(0, len(lines), batch_size)]

    # ── Parallel processing ───────────────────────────────────────────
    # Results indexed by batch_idx to maintain original line order
    batch_results: dict[int, list[dict]] = {}

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        MofNCompleteColumn(),
        console=console,
    ) as progress:
        task = progress.add_task(f"Classifying news ({len(batches)} batches)…", total=len(batches))

        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(
                    _classify_batch, idx, batch, system_prompt, client, config.model_name
                ): (idx, batch)
                for idx, batch in enumerate(batches)
            }

            for future in as_completed(futures):
                idx, batch = futures[future]
                batch_idx, records, err = future.result()

                if err:
                    progress.console.print(
                        f"[red]✖ Batch {batch_idx + 1} failed: {err}[/red]"
                    )
                    # Fallback: use raw text for failed batches
                    records = [
                        {
                            "topic": "Miscellaneous",
                            "tags": "",
                        }
                        for _ in batch
                    ]

                elif len(records) != len(batch):
                    progress.console.print(
                        f"[yellow]⚠ Batch {batch_idx + 1}: Expected {len(batch)} records, "
                        f"got {len(records)}. Padding/trimming.[/yellow]"
                    )
                    while len(records) < len(batch):
                        records.append({
                            "topic": "Miscellaneous",
                            "tags": "",
                        })
                    records = records[: len(batch)]
                else:
                    progress.console.print(
                        f"[green]✔ Batch {batch_idx + 1}: {len(records)} items classified[/green]"
                    )

                # Validate topics and inject month/year/news
                for i, rec in enumerate(records):
                    if rec.get("topic") not in STANDARD_TOPICS:
                        rec["topic"] = "Miscellaneous"
                    
                    # Inject CLI-provided and original line data
                    rec["month"] = month
                    rec["year"] = year
                    rec["news"] = batch[i].lstrip("- ").strip()

                batch_results[batch_idx] = records
                progress.advance(task)

    # ── Reassemble in original order ──────────────────────────────────
    all_records: list[dict] = []
    for idx in range(len(batches)):
        all_records.extend(batch_results.get(idx, []))

    # ── Write / append Excel ──────────────────────────────────────────
    console.print(f"\n[cyan]Writing {len(all_records)} records to Excel…[/cyan]")
    write_excel(all_records, output, source_filename=file_path.name)

    print_success(f"Done! {len(all_records)} news items → {output}")
    console.print("[dim]Open in Excel → use column header dropdowns to filter by Month, Year, Topic, Tags[/dim]")


@app.command("from-json")
def from_json(
    file_path: Path = typer.Argument(
        ...,
        exists=True,
        file_okay=True,
        dir_okay=False,
        help="Path to the JSON file containing current affairs data.",
    ),
    output: Path = typer.Option(
        None,
        "--output", "-o",
        help="Output Excel file path. Defaults to <input_name>.xlsx. Appends if exists.",
    )
):
    """
    Convert a generated JSON file directly into the filterable Excel spreadsheet.
    Requires no AI — purely deterministic parsing.
    """
    import json
    from aicli.services.news_parser import write_excel, STANDARD_TOPICS

    print_header(f"Parsing JSON: {file_path.name}")
    
    if output is None:
        output = file_path.parent / f"{file_path.stem}.xlsx"

    try:
        data = json.loads(file_path.read_text(encoding="utf-8"))
    except Exception as e:
        print_error("Failed to read JSON file.", e)
        raise typer.Exit(code=1)

    if not isinstance(data, list):
        print_error("JSON file must contain an array of objects.")
        raise typer.Exit(code=1)

    records = []
    for item in data:
        # Construct News string
        title = item.get("title", "").strip()
        key_answer = item.get("key_answer", "").strip()
        details = item.get("details", "").strip()
        
        news_parts = []
        if title:
            news_parts.append(f"**{title}**")
        if key_answer:
            news_parts.append(key_answer)
        if details:
            news_parts.append(details)
            
        news_str = "\n".join(news_parts)
        
        # Read keys (support upper and lower case added by user)
        source = item.get("Source") or item.get("source", "")
        order = item.get("Order") or item.get("order", "")
        
        # Standardize topic if possible
        topic = item.get("category", "")
        if topic not in STANDARD_TOPICS:
            if topic: 
                pass # keep whatever is in JSON if present
            else:
                topic = "Miscellaneous"

        src_val = str(source).strip() if source else ""
        ord_val = str(order).strip() if order else ""
        if src_val and ord_val:
            src_str = f"{src_val} - {ord_val}"
        else:
            src_str = src_val or ord_val
            
        records.append({
            "date": f"{item.get('month', 'Not Specified')} - {item.get('year', 'Not Specified')}",
            "topic": topic,
            "tags": str(item.get("tags", "")),
            "news": news_str,
            "source_key": src_str,
            "_raw_order": ord_val,
            "concat": ""
        })

    if output.exists():
        console.print(f"[yellow]⚡ Appending {len(records)} records to existing file:[/yellow] {output}")
    else:
        console.print(f"[green]Creating new file with {len(records)} records:[/green] {output}")

    write_excel(records, output, source_filename="")
    print_success(f"Done! {len(records)} items written to {output}")


@app.command("dedupe")
def dedupe(
    file_path: Path = typer.Argument(
        ...,
        exists=True,
        file_okay=True,
        dir_okay=False,
        help="Path to the Excel file to deduplicate.",
    ),
    output: Path = typer.Option(
        None,
        "--output", "-o",
        help="Output Excel file path. Defaults to <input_name>_deduped.xlsx.",
    ),
    threshold: float = typer.Option(
        0.80,
        "--threshold", "-t",
        help="Cosine similarity threshold for embeddings (0.0 to 1.0). Default is 0.80.",
    ),
    workers: int = typer.Option(
        10,
        "--workers", "-w",
        help="Number of parallel LLM workers for merging duplicates. Default is 10.",
    ),
):
    """
    De-duplicate an existing Current Affairs Excel file using RAG AI.
    
    1. Computes local deep semantic embeddings for every row using sentence-transformers.
    2. Clusters similar rows together via Cosine Similarity.
    3. Calls your local LLM to perfectly merge the duplicate news entries without data loss.
    """
    import openpyxl
    from openai import OpenAI
    from sentence_transformers import SentenceTransformer, util
    
    from aicli.config import config
    from aicli.services.news_parser import write_excel, merge_duplicate_news
    
    print_header(f"AI De-duplication: {file_path.name}")
    
    if output is None:
        output = file_path.parent / f"{file_path.stem}_deduped.xlsx"
        
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        print_error("Failed to read Excel file.", e)
        raise typer.Exit(code=1)
        
    # ── Smart Header Mapping ─────────────────────────────────────────
    # Identify column indices based on header names (case-insensitive)
    headers = [str(cell).strip().lower() for cell in next(ws.iter_rows(max_row=1, values_only=True))]
    
    def get_idx(candidates: list[str]) -> int:
        for c in candidates:
            if c.lower() in headers:
                return headers.index(c.lower())
        return -1

    idx_month = get_idx(["month"])
    idx_year = get_idx(["year"])
    idx_date = get_idx(["date", "month - year"])
    idx_topic = get_idx(["topic", "category"])
    idx_tags = get_idx(["tags"])
    idx_news = get_idx(["news", "details"])
    idx_source = get_idx(["source", "provider"])
    idx_order = get_idx(["order", "id"])
    idx_concat = get_idx(["concat", "original", "raw"])

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
            
        # Synthesize Date from Month/Year if needed
        month = str(row[idx_month]) if idx_month != -1 and row[idx_month] is not None else ""
        year = str(row[idx_year]) if idx_year != -1 and row[idx_year] is not None else ""
        
        if idx_date != -1 and row[idx_date]:
            date_val = str(row[idx_date])
        elif month or year:
            date_val = f"{month or 'Not Specified'} - {year or 'Not Specified'}"
        else:
            date_val = "Not Specified"

        source_val = str(row[idx_source]).strip() if idx_source != -1 and row[idx_source] is not None else ""
        order_val = str(row[idx_order]).strip() if idx_order != -1 and row[idx_order] is not None else ""
        if source_val and order_val:
            src_str = f"{source_val} - {order_val}"
        else:
            src_str = source_val or order_val
            
        records.append({
            "date": date_val,
            "topic": str(row[idx_topic]) if idx_topic != -1 and row[idx_topic] is not None else "Miscellaneous",
            "tags": str(row[idx_tags]) if idx_tags != -1 and row[idx_tags] is not None else "",
            "news": str(row[idx_news]) if idx_news != -1 and row[idx_news] is not None else "",
            "source_key": src_str,
            "_raw_order": order_val,
            "concat": str(row[idx_concat]) if idx_concat != -1 and row[idx_concat] is not None else "",
            # Keep raw parts for sorting if they exist
            "raw_month": month,
            "raw_year": year
        })
        
    if not records:
        print_error("No records found in Excel.")
        raise typer.Exit(code=1)

    # ── 1. Embeddings ────────────────────────────────────────────────
    console.print(f"[cyan]Loading local embedding model ('all-MiniLM-L6-v2')...[/cyan]")
    model = SentenceTransformer('all-MiniLM-L6-v2')
    
    news_texts = [r["news"] for r in records]
    console.print(f"[cyan]Computing semantic embeddings for {len(records)} records...[/cyan]")
    embeddings = model.encode(news_texts, convert_to_tensor=True)
    
    # ── 2. Clustering ────────────────────────────────────────────────
    console.print(f"[cyan]Clustering duplicates (Threshold > {threshold})...[/cyan]")
    cos_scores = util.cos_sim(embeddings, embeddings)
    
    visited = set()
    clusters = []
    
    for i in range(len(records)):
        if i in visited: continue
        cluster = [i]
        visited.add(i)
        for j in range(i + 1, len(records)):
            if j not in visited and cos_scores[i][j].item() >= threshold:
                cluster.append(j)
                visited.add(j)
        clusters.append(cluster)
        
    num_duplicates = sum(len(c) - 1 for c in clusters)
    if num_duplicates == 0:
        console.print("[green]✔ No duplicates found! Excel is perfectly clean.[/green]")
        raise typer.Exit()
        
    console.print(f"[yellow]⚠ Found {len(clusters)} unique events. {num_duplicates} duplicate records will be merged.[/yellow]")

    # ── 3. AI Merging ────────────────────────────────────────────────
    client = OpenAI(
        base_url=config.lm_studio_base_url,
        api_key=config.lm_studio_api_key,
    )
    
    unique_records = []
    
    # Separate pass-throughs from actual merges
    merge_clusters = [c for c in clusters if len(c) > 1]
    single_clusters = [c for c in clusters if len(c) == 1]
    
    # Pass-throughs: no LLM needed
    for cluster in single_clusters:
        unique_records.append(records[cluster[0]])
    
    # ── Pre-compute metadata for all merge clusters (fast, no LLM) ──
    merge_jobs = []
    for cluster in merge_clusters:
        items_to_merge = [records[idx] for idx in cluster]
        
        t_tags = []
        month = ""
        year = ""
        topic = "Miscellaneous"
        date_from_cluster = ""
        
        pairs = []
        seen = set()
        
        for item in items_to_merge:
            t_tags.extend([t.strip() for t in item["tags"].split(",") if t.strip()])
            
            if item.get("raw_month") and item["raw_month"] != "Not Specified": month = item["raw_month"]
            if item.get("raw_year") and item["raw_year"] != "Not Specified": year = item["raw_year"]
            if item.get("topic") and item["topic"] != "Miscellaneous": topic = item["topic"]
            if item.get("date") and item["date"] != "Not Specified": date_from_cluster = item["date"]
            
            sp = [s.strip() for s in item.get("source_key", "").split("|")]
            op = [o.strip() for o in item.get("order_key", "").split("|")]
            max_l = max(len(sp), len(op))
            sp += [""] * (max_l - len(sp))
            op += [""] * (max_l - len(op))
            for s, o in zip(sp, op):
                if (s, o) not in seen and (s or o):
                    seen.add((s, o))
                    pairs.append((s, o))

        unique_tags = list(dict.fromkeys(t_tags))
        merged_source_key = " | ".join(f"{p[0]} - {p[1]}" for p in pairs)
        raw_first_order = pairs[0][1] if pairs else ""
        
        if date_from_cluster:
            merged_date = date_from_cluster
        elif month or year:
            merged_date = f"{month or 'Not Specified'} - {year or 'Not Specified'}"
        else:
            merged_date = "Not Specified"

        news_strings = [i["news"] for i in items_to_merge]
        visual_concat = "\n\n---\n".join([f"RECORD {i}:\n{item}" for i, item in enumerate(news_strings, 1)])
        
        merge_jobs.append({
            "date": merged_date,
            "topic": topic,
            "tags": ", ".join(unique_tags),
            "source_key": merged_source_key,
            "_raw_order": raw_first_order,
            "concat": visual_concat,
            "_news_strings": news_strings,
        })

    # ── Parallel LLM merging ─────────────────────────────────────────
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    
    save_lock = threading.Lock()
    
    def _save_snapshot(recs, path):
        """Thread-safe progressive save — writes current state to Excel."""
        with save_lock:
            if path.exists():
                path.unlink()
            write_excel(list(recs), path, source_filename="")
    
    def _do_merge(job):
        news_strs = job.pop("_news_strings")
        merged = merge_duplicate_news(news_strs, client, config.model_name)
        job["news"] = merged
        return job
    
    console.print(f"[cyan]  Launching {workers} parallel LLM workers for {len(merge_jobs)} merge tasks...[/cyan]")
    
    # Must remove the output file if it exists, because write_excel appends
    if output.exists():
        output.unlink()
    
    # Save pass-throughs immediately as a baseline snapshot
    _save_snapshot(unique_records, output)
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        MofNCompleteColumn(),
        console=console,
    ) as progress:
        task = progress.add_task(f"Merging {len(merge_jobs)} duplicates...", total=len(merge_jobs))
        
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(_do_merge, job): i for i, job in enumerate(merge_jobs)}
            for future in as_completed(futures):
                result = future.result()
                unique_records.append(result)
                progress.advance(task)
                
                # ── Save on the go: write after each merge ──
                _save_snapshot(unique_records, output)

    def sort_key(rec):
        # Sort by first source alphabetically, then by raw order numerically
        src = rec["source_key"].split("|")[0].split("-")[0].strip().lower()
        ord_val = rec.get("_raw_order", "").split("|")[0].strip()
            
        try:
            return (src, int(ord_val))
        except ValueError:
            return (src, 999999)

    unique_records.sort(key=sort_key)

    console.print(f"[green]✔ AI De-duplication complete. Merged {num_duplicates} duplicate records.[/green]")
    console.print(f"[cyan]Writing {len(unique_records)} sorted, pristine records to Excel...[/cyan]")
    
    # Final sorted save
    if output.exists():
        output.unlink()
    write_excel(unique_records, output, source_filename="")
    print_success(f"Deduplicated file saved → {output}")

@app.command("process")
def process_news(
    json_path: Path = typer.Argument(
        ...,
        exists=True,
        file_okay=True,
        dir_okay=False,
        help="Path to the JSON file containing new current affairs data.",
    ),
    output: Path = typer.Option(
        None,
        "--output", "-o",
        help="Master Output Excel file path. Defaults to <input_name>_master.xlsx.",
    ),
    workers: int = typer.Option(
        4,
        "--workers",
        "-w",
        help="Number of parallel LLM inference threads.",
        min=1,
    ),
    threshold: float = typer.Option(
        0.8,
        "--threshold",
        "-t",
        help="Cosine similarity threshold for duplicates (0.0 to 1.0).",
    )
):
    """
    Unified God-Mode Pipeline: Parses JSON, appends to existing master Excel, and natively deduplicates the entire dataset.
    """
    import json
    import openpyxl
    import threading
    from sentence_transformers import SentenceTransformer, util
    from openai import OpenAI
    from aicli.config import config
    from aicli.services.news_parser import write_excel, STANDARD_TOPICS, build_system_prompt, merge_duplicate_news
    
    # ── 1. Parse JSON block (New Data) ─────────────────────────────────
    print_header("God-Mode Processing")
    console.print(f"[cyan]Parsing input JSON: {json_path.name}...[/cyan]")
    
    if output is None:
        output = json_path.parent / f"{json_path.stem}_master.xlsx"
        
    try:
        data = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as e:
        print_error("Failed to read JSON file.", e)
        raise typer.Exit(code=1)

    if not isinstance(data, list):
        print_error("JSON file must contain an array of objects.")
        raise typer.Exit(code=1)

    records = []
    for item in data:
        title = item.get("title", "").strip()
        key_answer = item.get("key_answer", "").strip()
        details = item.get("details", "").strip()
        
        news_parts = []
        if title:
            news_parts.append(f"**{title}**")
        if key_answer:
            news_parts.append(key_answer)
        if details:
            news_parts.append(details)
            
        news_str = "\n".join(news_parts)
        
        source = item.get("Source") or item.get("source", "")
        order = item.get("Order") or item.get("order", "")
        
        topic = item.get("category", "")
        if topic not in STANDARD_TOPICS:
            if topic: 
                pass 
            else:
                topic = "Miscellaneous"

        src_val = str(source).strip() if source else ""
        ord_val = str(order).strip() if order else ""
        if src_val and ord_val:
            src_str = f"{src_val} - {ord_val}"
        else:
            src_str = src_val or ord_val
            
        records.append({
            "date": f"{item.get('month', 'Not Specified')} - {item.get('year', 'Not Specified')}",
            "topic": topic,
            "tags": str(item.get("tags", "")),
            "news": news_str,
            "source_key": src_str,
            "_raw_order": ord_val,
            "concat": ""  # Handled below
        })
        
    console.print(f"[green]✔ Extracted {len(records)} new records from JSON.[/green]")
    
    # ── 2. Load Existing Master Excel ──────────────────────────────────
    if output.exists():
        console.print(f"[cyan]Master database exists. Loading existing records from {output.name}...[/cyan]")
        try:
            wb = openpyxl.load_workbook(output)
            ws = wb.active
            
            headers = [str(cell).strip().lower() for cell in next(ws.iter_rows(max_row=1, values_only=True))]
            def get_idx(candidates: list[str]) -> int:
                for c in candidates:
                    if c.lower() in headers:
                        return headers.index(c.lower())
                return -1

            idx_month = get_idx(["month"]) # Fallback if no date 
            idx_year = get_idx(["year"])   # Fallback if no date
            idx_date = get_idx(["date", "month - year"])
            idx_topic = get_idx(["topic", "category"])
            idx_tags = get_idx(["tags"])
            idx_news = get_idx(["news", "details"])
            idx_source = get_idx(["source", "provider"])
            idx_order = get_idx(["order", "id"])
            idx_concat = get_idx(["concat", "original", "raw"])
            
            existing_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                
                month = str(row[idx_month]) if idx_month != -1 and row[idx_month] is not None else ""
                year = str(row[idx_year]) if idx_year != -1 and row[idx_year] is not None else ""
                if idx_date != -1 and row[idx_date]:
                    date_val = str(row[idx_date])
                elif month or year:
                    date_val = f"{month or 'Not Specified'} - {year or 'Not Specified'}"
                else:
                    date_val = "Not Specified"
                    
                s_val = str(row[idx_source]).strip() if idx_source != -1 and row[idx_source] is not None else ""
                o_val = str(row[idx_order]).strip() if idx_order != -1 and row[idx_order] is not None else ""
                if s_val and o_val:
                    s_str = f"{s_val} - {o_val}"
                else:
                    s_str = s_val or o_val
                    
                records.append({
                    "date": date_val,
                    "topic": str(row[idx_topic]) if idx_topic != -1 and row[idx_topic] is not None else "Miscellaneous",
                    "tags": str(row[idx_tags]) if idx_tags != -1 and row[idx_tags] is not None else "",
                    "news": str(row[idx_news]) if idx_news != -1 and row[idx_news] is not None else "",
                    "source_key": s_str,
                    "_raw_order": o_val,
                    "concat": str(row[idx_concat]) if idx_concat != -1 and row[idx_concat] is not None else "",
                    "raw_month": month,
                    "raw_year": year
                })
                existing_count += 1
            console.print(f"[green]✔ Loaded {existing_count} existing records. Total pool: {len(records)} items.[/green]")
        except Exception as e:
            print_error("Failed to read existing Excel file.", e)
            raise typer.Exit(code=1)
    else:
        console.print(f"[yellow]Master database does not exist. A new one will be created: {output.name}[/yellow]")
        
    # ── 3. Embeddings & Clustering ─────────────────────────────────────
    console.print(f"\n[cyan]Loading local embedding model ('all-MiniLM-L6-v2')...[/cyan]")
    model = SentenceTransformer('all-MiniLM-L6-v2')
    
    news_texts = [r["news"] for r in records]
    console.print(f"[cyan]Computing semantic embeddings for {len(records)} records...[/cyan]")
    embeddings = model.encode(news_texts, convert_to_tensor=True)
    
    console.print(f"[cyan]Clustering duplicates (Threshold > {threshold})...[/cyan]")
    cos_scores = util.cos_sim(embeddings, embeddings)
    
    visited = set()
    clusters = []
    
    for i in range(len(records)):
        if i in visited: continue
        cluster = [i]
        visited.add(i)
        for j in range(i + 1, len(records)):
            if j not in visited and cos_scores[i][j].item() >= threshold:
                cluster.append(j)
                visited.add(j)
        clusters.append(cluster)
        
    num_duplicates = sum(len(c) - 1 for c in clusters)
    if num_duplicates == 0:
        console.print("[green]✔ No duplicates found! Pool is perfectly clean.[/green]")
        if output.exists():
            output.unlink()
        write_excel(records, output, source_filename="")
        print_success(f"File updated and saved → {output}")
        raise typer.Exit()
        
    console.print(f"[yellow]⚠ Found {len(clusters)} unique events. {num_duplicates} duplicate records will be merged.[/yellow]")

    # ── 4. AI Merging with Progressive Saves ───────────────────────────
    client = OpenAI(
        base_url=config.lm_studio_base_url,
        api_key=config.lm_studio_api_key,
    )
    
    unique_records = []
    merge_clusters = [c for c in clusters if len(c) > 1]
    single_clusters = [c for c in clusters if len(c) == 1]
    
    # Pass-throughs
    for cluster in single_clusters:
        unique_records.append(records[cluster[0]])
        
    # Pre-compute metadata for merge clusters
    merge_jobs = []
    for cluster in merge_clusters:
        items_to_merge = [records[idx] for idx in cluster]
        
        t_tags = []
        month = ""
        year = ""
        topic = "Miscellaneous"
        date_from_cluster = ""
        
        pairs = []
        seen = set()
        
        for item in items_to_merge:
            t_tags.extend([t.strip() for t in item["tags"].split(",") if t.strip()])
            
            if item.get("raw_month") and item["raw_month"] != "Not Specified": month = item["raw_month"]
            if item.get("raw_year") and item["raw_year"] != "Not Specified": year = item["raw_year"]
            if item.get("topic") and item["topic"] != "Miscellaneous": topic = item["topic"]
            if item.get("date") and item["date"] != "Not Specified": date_from_cluster = item["date"]
            
            sp = [s.strip() for s in item.get("source_key", "").split("|")]
            op = [o.strip() for o in item.get("_raw_order", "").split("|")]
            max_l = max(len(sp), len(op))
            sp += [""] * (max_l - len(sp))
            op += [""] * (max_l - len(op))
            for s, o in zip(sp, op):
                if (s, o) not in seen and (s or o):
                    seen.add((s, o))
                    pairs.append((s, o))

        unique_tags = list(dict.fromkeys(t_tags))
        merged_source_key = " | ".join(f"{p[0]} - {p[1]}" for p in pairs)
        raw_first_order = pairs[0][1] if pairs else ""
        
        if date_from_cluster:
            merged_date = date_from_cluster
        elif month or year:
            merged_date = f"{month or 'Not Specified'} - {year or 'Not Specified'}"
        else:
            merged_date = "Not Specified"

        news_strings = [i["news"] for i in items_to_merge]
        visual_concat = "\n\n---\n".join([f"RECORD {i}:\n{item}" for i, item in enumerate(news_strings, 1)])
        
        merge_jobs.append({
            "date": merged_date,
            "topic": topic,
            "tags": ", ".join(unique_tags),
            "source_key": merged_source_key,
            "_raw_order": raw_first_order,
            "concat": visual_concat,
            "_news_strings": news_strings,
        })
    
    file_lock = threading.Lock()
    def _save_snapshot(current_merged, output_file):
        with file_lock:
            if output_file.exists():
                output_file.unlink()
            write_excel(current_merged, output_file, source_filename="")
            
    def _do_merge(job):
        news_strs = job.pop("_news_strings")
        # Check if already merged (progressive bypass is nice, but we lack the 'unmerged_new' array in this refactor. Let's just merge all strings natively.)
        merged = merge_duplicate_news(news_strs, client, config.model_name)
        job["news"] = merged
        return job

    console.print(f"  [dim]Launching {workers} parallel LLM workers for {len(merge_jobs)} merge tasks...[/dim]")
    
    # Initial save snapshot with passthroughs
    _save_snapshot(unique_records, output)
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        MofNCompleteColumn(),
        console=console,
    ) as progress:
        task = progress.add_task(f"Merging {len(merge_jobs)} duplicates...", total=len(merge_jobs))
        
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(_do_merge, job): i for i, job in enumerate(merge_jobs)}
            for future in as_completed(futures):
                result = future.result()
                unique_records.append(result)
                progress.advance(task)
                
                _save_snapshot(unique_records, output)

    def sort_key(rec):
        src = rec.get("source_key", "").split("|")[0].split("-")[0].strip().lower()
        ord_val = rec.get("_raw_order", "").split("|")[0].strip()
        try:
            return (src, int(ord_val))
        except ValueError:
            return (src, 999999)

    unique_records.sort(key=sort_key)

    console.print(f"[green]✔ AI De-duplication complete. Merged {num_duplicates} duplicate records.[/green]")
    console.print(f"[cyan]Writing {len(unique_records)} sorted, pristine records to Excel...[/cyan]")
    
    if output.exists():
        output.unlink()
    write_excel(unique_records, output, source_filename="")
    print_success(f"Master file updated and saved → {output}")
