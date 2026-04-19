"""
News Excel Repository
Handles all OpenPyxl formatting, dynamic column generation, reading, and writing for the filterable Current Affairs sheets.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class NewsExcelRepository:
    @staticmethod
    def _collect_unique_sources(records: list[dict]) -> list[str]:
        sources = set()
        for rec in records:
            src = rec.get("sources", {})
            if isinstance(src, dict):
                sources.update(src.keys())
            elif isinstance(src, str) and src:
                for part in src.split("|"):
                    if " - " in part:
                        sources.add(part.split(" - ")[0].strip())
                    else:
                        sources.add(part.strip())
        return sorted(sources)

    @staticmethod
    def _build_dynamic_headers(sources: list[str]) -> list[str]:
        return ["S.No", "Date", "Topic", "Tags", "News"] + sources + ["Concat"]

    @staticmethod
    def _get_col_widths_dynamic(num_sources: int) -> dict:
        widths = {"A": 6, "B": 18, "C": 30, "D": 36, "E": 80}
        source_cols = "FGHIJKLMNOPQRSTUVWXYZ"
        for i in range(num_sources):
            if i < len(source_cols):
                widths[source_cols[i]] = 12
        widths[chr(ord("A") + 5 + num_sources)] = 60
        return widths

    @staticmethod
    def _style_header_row(ws, headers: list[str]) -> None:
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="8DB4E2"),
            right=Side(style="thin", color="8DB4E2"),
            top=Side(style="thin", color="8DB4E2"),
            bottom=Side(style="medium", color="1F3864"),
        )
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    @staticmethod
    def _style_data_row(ws, row_num: int, values: list, row_index: int) -> None:
        data_font = Font(name="Calibri", size=10)
        even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="B4C6E7"),
            right=Side(style="thin", color="B4C6E7"),
            top=Side(style="thin", color="B4C6E7"),
            bottom=Side(style="thin", color="B4C6E7"),
        )

        fill = even_fill if row_index % 2 == 0 else odd_fill

        news_col_idx = 5
        news_text = values[news_col_idx - 1] if news_col_idx <= len(values) else ""
        line_count = news_text.count("\n") + 1
        row_height = max(15, min(line_count * 12, 300))

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            if col in (1, 2):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.row_dimensions[row_num].height = row_height

    @staticmethod
    def _apply_finishing(ws, total_rows: int, col_widths: dict) -> None:
        last_col_letter = list(col_widths.keys())[-1]
        ws.auto_filter.ref = f"A1:{last_col_letter}{total_rows}"
        ws.freeze_panes = "A2"
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

    @classmethod
    def write_excel(cls, records: list[dict], output_path: Path) -> None:
        all_records = []
        if output_path.exists():
            all_records = cls.read_excel_records(output_path)

        all_records.extend(records)

        unique_sources = cls._collect_unique_sources(all_records)
        headers = cls._build_dynamic_headers(unique_sources)
        col_widths = cls._get_col_widths_dynamic(len(unique_sources))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Current Affairs"
        cls._style_header_row(ws, headers)

        for idx, record in enumerate(records):
            global_idx = idx
            serial = idx + 1
            row_num = serial + 1

            sources_dict = record.get("sources", {})
            if isinstance(sources_dict, str):
                src_dict = {}
                for part in sources_dict.split("|"):
                    if " - " in part:
                        name, order = part.split(" - ", 1)
                        src_dict[name.strip()] = order.strip()
                sources_dict = src_dict

            source_values = [str(sources_dict.get(src, "")) for src in unique_sources]
            concat_value = record.get("concat_json", "")

            values = (
                [
                    serial,
                    record.get("date", "Not Specified"),
                    record.get("topic", "Miscellaneous"),
                    record.get("tags", ""),
                    record.get("news", ""),
                ]
                + source_values
                + [concat_value]
            )
            cls._style_data_row(ws, row_num, values, global_idx)

        total_rows = len(records) + 1
        cls._apply_finishing(ws, total_rows, col_widths)

        wb.save(output_path)

    @staticmethod
    def read_excel_records(path: Path) -> list[dict]:
        if not path.exists():
            return []
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [str(cell).strip() for cell in next(ws.iter_rows(max_row=1, values_only=True))]

        known_cols = {"s.no", "date", "topic", "tags", "news", "concat"}
        source_cols = [h for h in headers if h.lower() not in known_cols and h.lower() != "order"]

        indices = {h.lower(): i for i, h in enumerate(headers)}
        idx_date = indices.get("date", -1)
        idx_topic = indices.get("topic", -1)
        idx_tags = indices.get("tags", -1)
        idx_news = indices.get("news", -1)
        idx_concat = indices.get("concat", -1)

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            sources = {}
            for src_col in source_cols:
                col_idx = headers.index(src_col)
                val = row[col_idx] if col_idx < len(row) else None
                if val:
                    sources[src_col] = str(val)

            concat_val = row[idx_concat] if idx_concat != -1 and idx_concat < len(row) else ""

            records.append({
                "date": str(row[idx_date]) if idx_date != -1 and idx_date < len(row) else "Not Specified",
                "topic": str(row[idx_topic]) if idx_topic != -1 and idx_topic < len(row) else "Miscellaneous",
                "tags": str(row[idx_tags]) if idx_tags != -1 and idx_tags < len(row) else "",
                "news": str(row[idx_news]) if idx_news != -1 and idx_news < len(row) else "",
                "sources": sources,
                "concat_json": str(concat_val) if concat_val else "",
            })
        return records
