"""
News Parser Service — Current Affairs → Filterable Excel.

Reads a text file where each line is one news item, classifies each via LLM
into (Month, Year, Topic, Tags, News), and writes a styled, filterable Excel file.
Supports APPEND mode so multiple source files accumulate into one master sheet.
"""

import json
import re
from pathlib import Path


# ──────────────────────────────────────────────────────────────────────
# Standardized Topic List (competitive‑exam categories)
# ──────────────────────────────────────────────────────────────────────

STANDARD_TOPICS: list[str] = [
    "Polity & Governance",
    "International Relations",
    "Defence & Security",
    "Economy & Finance",
    "Science & Technology",
    "Environment & Ecology",
    "Art, Culture & Heritage",
    "Awards & Honours",
    "Sports",
    "Appointments & Personnel",
    "Books & Authors",
    "Important Days & Dates",
    "Government Schemes & Programs",
    "Summits & Conferences",
    "Reports & Indices",
    "Obituaries",
    "Infrastructure & Energy",
    "Space & Exploration",
    "Health & Medicine",
    "Education",
    "Law & Judiciary",
    "Miscellaneous",
]


# ──────────────────────────────────────────────────────────────────────
# LLM Prompts
# ──────────────────────────────────────────────────────────────────────


def build_system_prompt() -> str:
    """Build the system prompt with the topic list baked in."""
    topics_block = "\n".join(f"  - {t}" for t in STANDARD_TOPICS)

    return f"""You are a precise Current Affairs classification engine for competitive exam preparation (UPSC/SSC/Banking).

You will receive one or more news items, each on its own numbered line. For EACH line, extract structured data and return it as a JSON array.

## OUTPUT FORMAT
Return a JSON array. Each element corresponds to one input line IN ORDER and must have exactly these 2 keys:
[
  {{
    "topic": "<Exactly one topic from the STANDARD LIST below>",
    "tags": "<Comma-separated micro-level keywords for granular filtering>"
  }}
]

## STANDARD TOPICS — USE EXACTLY ONE PER LINE
{topics_block}

## TOPIC CLASSIFICATION GUIDELINES
- Military exercises, missiles, warships, defense exports, military operations → "Defence & Security"
- Bilateral/multilateral relations, foreign visits, treaties, diplomatic agreements → "International Relations"
- UNESCO heritage, festivals, GI tags, traditional art forms, cultural events → "Art, Culture & Heritage"
- Person died / passed away / demise → "Obituaries"
- Person appointed / elected / nominated to a position → "Appointments & Personnel"
- Book launched / authored / released → "Books & Authors"
- Named day observances (World AIDS Day, Human Rights Day, etc.) → "Important Days & Dates"
- Government bills, acts, schemes, welfare missions, policy launches → "Government Schemes & Programs"
- Repo rate, GDP, UPI, financial systems, stock markets, budgets → "Economy & Finance"
- Satellites, rockets, ISRO/NASA space missions, astronauts → "Space & Exploration"
- Rankings, global indices, survey reports → "Reports & Indices"
- Sports tournaments, sports results, player records, auction → "Sports"
- Awards/prizes NOT sport-tournament results → "Awards & Honours"
- Dams, power plants, highways, airports, ports, energy projects → "Infrastructure & Energy"
- Biodiversity, climate, wildlife, conservation, pollution, Ramsar sites → "Environment & Ecology"
- Constitutional matters, governance reforms, renaming of institutions → "Polity & Governance"
- Disease control, WHO health declarations, medical breakthroughs → "Health & Medicine"
- Universities, education boards, academic events → "Education"
- Court judgments, impeachment, legal proceedings → "Law & Judiciary"
- COP/biodiversity/health summits, bilateral summits, international conferences → "Summits & Conferences"
- If genuinely ambiguous, use "Miscellaneous"

## TAGS RULES
- Tags are flexible, micro-level keywords that are MORE GRANULAR than the broad Topic.
- They help with fine-grained filtering beyond the standardized topic.
- Provide 2-5 comma-separated tags per news item.
- Tags should include: key entity names, geographic locations, organizations, specific sub-domains.
- Examples:
  - Topic "Defence & Security", Tags: "DRDO, Missile, K4, Nuclear, INS Arighat"
  - Topic "Sports", Tags: "Cricket, IPL, Auction, Cameron Green"
  - Topic "Environment & Ecology", Tags: "Ramsar, Wetland, Rajasthan, Siliserh Lake"
  - Topic "Art, Culture & Heritage", Tags: "UNESCO, Diwali, Intangible Heritage"
  - Topic "International Relations", Tags: "India, Russia, Summit, Putin, Delhi"
- Tags should be proper nouns or specific terms — NOT generic adjectives.

## STRICT OUTPUT RULES
- Output ONLY the raw JSON array. No markdown code fences, no explanation, no commentary.
- Every element must have all 2 keys.
- The array must have EXACTLY as many elements as input lines, in the same order."""


def build_user_prompt(lines: list[str]) -> str:
    """Build the user prompt with numbered lines."""
    numbered = "\n".join(f"{i + 1}. {line}" for i, line in enumerate(lines))
    return (
        f"Classify each of the following {len(lines)} news line(s). "
        f"Return a JSON array with exactly {len(lines)} element(s), one per line, in the same order.\n\n"
        f"{numbered}"
    )


# ──────────────────────────────────────────────────────────────────────
# Parsing helpers
# ──────────────────────────────────────────────────────────────────────


def extract_json_array(raw: str) -> list[dict]:
    """Robustly extract a JSON array from potentially noisy LLM output."""
    text = raw.strip()

    # Strip markdown code fences
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    text = text.strip()

    # Attempt 1: direct parse
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            return parsed
    except json.JSONDecodeError:
        pass

    # Attempt 2: find the largest [...] block
    match = re.search(r"\[.*\]", text, re.DOTALL)
    if match:
        try:
            parsed = json.loads(match.group())
            if isinstance(parsed, list):
                return parsed
        except json.JSONDecodeError:
            pass

    return []


def filter_news_lines(filepath: Path) -> list[str]:
    """
    Read the file and return only lines that look like news items.
    Skips: empty lines and horizontal rules (---).
    """
    raw = filepath.read_text(encoding="utf-8")
    result: list[str] = []

    for line in raw.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if stripped == "---":
            continue
        result.append(stripped)

    return result


def merge_duplicate_news(news_items: list[str], client, model_name: str) -> str:
    """
    Given a list of semantically similar news strings, use the LLM to merge them into a single, comprehensive string.
    Ensures zero loss of factual detail.
    """
    from rich.console import Console

    console = Console()

    # Pre-concatenate the items in Python as requested
    raw_concatenation = "\n\n---\n".join(news_items)

    system_prompt = (
        "You are an expert editor handling current affairs data for Indian Competitive Exams.\n"
        "You will be given a concatenated block of duplicate news items describing the exact same event.\n"
        "Your task is to merge them into a single cohesive entry while strictly mimicking ultra-dense study notes.\n"
        "CRITICAL RULES (NON-NEGOTIABLE):\n"
        "1. FORMAT: Always start with a **Bold Title**. Use a double newline, then optionally a one-line summary, then the bulleted details.\n"
        "2. ZERO DATA LOSS: Every single fact, number, date, name, percentage, and detail from ALL source items MUST appear in your output. You are merging facts, NOT summarizing them away. Losing even one unique fact is UNACCEPTABLE.\n"
        "3. DETAILS FORMATTING (ULTRA-DENSE): Use **telegraphic language**. No 'is', 'was', 'the', or 'has been' unless necessary for meaning. Eliminate filler words and obvious explanations.\n"
        "4. STRUCTURE: Each distinct fact MUST be on its own line starting with a bullet point (e.g. `- **India's Rank**: 16th/154 countries.`).\n"
        "5. PRESERVE SPECIFICITY: Keep exact numbers, dates, percentages, names. Never round or paraphrase.\n"
        "6. Output ONLY the perfectly merged text block. Do NOT include markdown code fences (like ```), JSON, or conversational replies."
    )

    user_prompt = f"Merge and synthesize this concatenated news data into one clean record:\n\n{raw_concatenation}"

    import time

    max_retries = 5
    base_delay = 2.0

    for attempt in range(max_retries):
        try:
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.0,
            )
            content = response.choices[0].message.content.strip()
            return content

        except Exception as e:
            # If it's the last attempt, fail. Otherwise wait and retry.
            if attempt == max_retries - 1:
                console.print(
                    f"[red]⚠ AI Merge Error on cluster of size {len(news_items)} after {max_retries} attempts: {str(e)}[/red]"
                )
                return (
                    f"[MERGE ERROR - FALLBACK CONCATENATION]:\n\n" + raw_concatenation
                )

            # Check for model unloaded / context size issues which are concurrency symptoms
            err_msg = str(e).lower()
            if (
                "model unloaded" in err_msg
                or "context size" in err_msg
                or "no models loaded" in err_msg
                or "failed to parse" in err_msg
            ):
                sleep_time = base_delay * (2**attempt)
                time.sleep(sleep_time)
            else:
                # For other hard errs, just retry with standard backoff
                time.sleep(base_delay)


# ──────────────────────────────────────────────────────────────────────
# Excel writer — supports CREATE and APPEND
# ──────────────────────────────────────────────────────────────────────


def _collect_unique_sources(records: list[dict]) -> list[str]:
    """Extract all unique source names from records (supports both old and new format)."""
    sources = set()
    for rec in records:
        src = rec.get("sources", {})
        if isinstance(src, dict):
            sources.update(src.keys())
        elif isinstance(src, str) and src:
            for part in src.split("|"):
                if " - " in part:
                    sources.add(part.split(" - ")[0].strip())
                else:
                    sources.add(part.strip())
    return sorted(sources)


def _build_dynamic_headers(sources: list[str]) -> list[str]:
    """Build header row dynamically based on unique sources."""
    return ["S.No", "Date", "Topic", "Tags", "News"] + sources + ["Concat"]


def _get_col_widths_dynamic(num_sources: int) -> dict:
    """Build column widths dynamically based on number of sources."""
    widths = {"A": 6, "B": 18, "C": 30, "D": 36, "E": 80}
    source_cols = "FGHIJKLMNOPQRSTUVWXYZ"
    for i in range(num_sources):
        if i < len(source_cols):
            widths[source_cols[i]] = 12
    widths[chr(ord("A") + 5 + num_sources)] = 60
    return widths


def _style_header_row(ws, headers: list[str]) -> None:
    """Apply professional styling to the header row."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="1F3864", end_color="1F3864", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="8DB4E2"),
        right=Side(style="thin", color="8DB4E2"),
        top=Side(style="thin", color="8DB4E2"),
        bottom=Side(style="medium", color="1F3864"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _style_data_row(
    ws, row_num: int, values: list, row_index: int, num_sources: int
) -> None:
    """Apply styling to a single data row."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data_font = Font(name="Calibri", size=10)
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )

    fill = even_fill if row_index % 2 == 0 else odd_fill

    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.font = data_font
        cell.fill = fill
        cell.border = thin_border
        if col in (1, 2):  # S.No, Date — centered
            cell.alignment = Alignment(horizontal="center", vertical="top")
        else:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _apply_finishing(ws, total_rows: int, col_widths: dict) -> None:
    """Apply auto-filter, freeze panes, and column widths."""
    last_col_letter = list(col_widths.keys())[-1]
    ws.auto_filter.ref = f"A1:{last_col_letter}{total_rows}"
    ws.freeze_panes = "A2"
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width


def write_excel(records: list[dict], output_path: Path, source_filename: str) -> None:
    """
    Write or APPEND parsed news records to a filterable Excel file.

    - If `output_path` does not exist → creates a new workbook.
    - If `output_path` already exists → loads it, appends new rows, re-numbers S.No.
    - Dynamically creates source columns based on unique sources in records.
    """
    from openpyxl import Workbook, load_workbook

    all_records = []
    if output_path.exists():
        all_records = read_excel_records(output_path)

    all_records.extend(records)

    unique_sources = _collect_unique_sources(all_records)
    headers = _build_dynamic_headers(unique_sources)
    col_widths = _get_col_widths_dynamic(len(unique_sources))

    wb = Workbook()
    ws = wb.active
    ws.title = "Current Affairs"
    _style_header_row(ws, headers)

    for idx, record in enumerate(records):
        global_idx = idx
        serial = idx + 1
        row_num = serial + 1

        sources_dict = record.get("sources", {})
        if isinstance(sources_dict, str):
            src_dict = {}
            for part in sources_dict.split("|"):
                if " - " in part:
                    name, order = part.split(" - ", 1)
                    src_dict[name.strip()] = order.strip()
            sources_dict = src_dict

        source_values = [str(sources_dict.get(src, "")) for src in unique_sources]
        concat_value = record.get("concat_json", "")

        values = (
            [
                serial,
                record.get("date", "Not Specified"),
                record.get("topic", "Miscellaneous"),
                record.get("tags", ""),
                record.get("news", ""),
            ]
            + source_values
            + [concat_value]
        )
        _style_data_row(ws, row_num, values, global_idx, len(unique_sources))

    total_rows = len(records) + 1
    _apply_finishing(ws, total_rows, col_widths)

    wb.save(output_path)


def read_excel_records(path: Path) -> list[dict]:
    """Read records from Excel file, handling dynamic source columns."""
    import openpyxl

    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [
        str(cell).strip() for cell in next(ws.iter_rows(max_row=1, values_only=True))
    ]

    known_cols = {"s.no", "date", "topic", "tags", "news", "concat"}
    source_cols = [
        h for h in headers if h.lower() not in known_cols and h.lower() != "order"
    ]

    idx_sno = -1
    idx_date = -1
    idx_topic = -1
    idx_tags = -1
    idx_news = -1
    idx_concat = -1

    for i, h in enumerate(headers):
        h_lower = h.lower()
        if h_lower == "s.no":
            idx_sno = i
        elif h_lower == "date":
            idx_date = i
        elif h_lower == "topic":
            idx_topic = i
        elif h_lower == "tags":
            idx_tags = i
        elif h_lower == "news":
            idx_news = i
        elif h_lower == "concat":
            idx_concat = i

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        sources = {}
        for src_col in source_cols:
            col_idx = headers.index(src_col)
            val = row[col_idx] if col_idx < len(row) else None
            if val:
                sources[src_col] = str(val)

        concat_val = (
            row[idx_concat] if idx_concat != -1 and idx_concat < len(row) else ""
        )

        records.append(
            {
                "date": str(row[idx_date])
                if idx_date != -1 and idx_date < len(row)
                else "Not Specified",
                "topic": str(row[idx_topic])
                if idx_topic != -1 and idx_topic < len(row)
                else "Miscellaneous",
                "tags": str(row[idx_tags])
                if idx_tags != -1 and idx_tags < len(row)
                else "",
                "news": str(row[idx_news])
                if idx_news != -1 and idx_news < len(row)
                else "",
                "sources": sources,
                "concat_json": str(concat_val) if concat_val else "",
            }
        )
    return records
